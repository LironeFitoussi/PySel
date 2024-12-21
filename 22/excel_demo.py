import openpyxl
import os
# Load the workbook
book = openpyxl.load_workbook('./PythonDemo.xlsx')

# Select the active sheet
sheet = book.active

Dict = {}
# Example: Read a cell value
# cell_value = sheet['A1'].value
# print(f'The value of cell A1 is: {cell_value}')

# print(sheet.max_row)
# Example: Write a value to a cell
# sheet['A5'] = 'Hello, World!'

for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value == 'Test Case 1':
        for col in range(1, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value 
            # Dict = {'Test Name': 'Test Case 1', 'first_name': 'Lirone', 'last_name': 'Fitoussi', 'password': 123456789, 'gender': 'Male'}


print(Dict)

# Save the workbooks
book.save('./PythonDemo.xlsx')

# Open the workbook in Excel to see the changes (Mac)
# os.system('open ./PythonDemo.xlsx')
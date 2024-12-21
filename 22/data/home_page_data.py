import openpyxl

class HomePageData:
    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook("./PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == test_case_name:
                for col in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=col).value] = (
                        sheet.cell(row=row, column=col).value
                    )
        
        return [Dict]

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
import os

# Path to the file
file_path = '/Users/lironefitoussi/Downloads/download.xlsx'
toast_locator = (By.CSS_SELECTOR, '.Toastify__toast-body div:nth-child(2)')
fruit = 'Apple'
new_price = 350

def edit_excel(file_path, fruit_name, column, new_price):
    # Edit the Excel file and change the price of the fruit
    if os.path.exists(file_path):
        # 1. Open the Excel file
        book = openpyxl.load_workbook(file_path)
        sheet = book.active

        # Get Row Count
        row_count = sheet.max_row
        
        # 2. Find the row with the fruit and update the price
        for row in range(1, row_count + 1):
            if sheet.cell(row=row, column=2).value == fruit_name:
                sheet.cell(row=row, column=column).value = new_price
                break

        # 3. Save the file
        book.save(file_path)
    else:
        print("File not found. Exiting.")
        driver.quit()
        exit()

# Browser setup
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get('http://rahulshettyacademy.com/upload-download-test/index.html')

wait = WebDriverWait(driver, 10)

# Click on the Download button
downloadBtn = driver.find_element(By.ID, 'downloadButton')
downloadBtn.click()

# Wait for the file to download
time.sleep(5)  # Replace with proper file-checking logic if needed

# Edit the Excel file
edit_excel(file_path, fruit, 4, new_price)

# Upload the edited Excel file
uploadBtn = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
uploadBtn.send_keys(file_path)

# Wait for the toast notification
element = wait.until(EC.visibility_of_element_located(toast_locator))
print(element.text)

# Delete the file
if os.path.exists(file_path):
    os.remove(file_path)

# Check if the price has changed on the webpage
fruit_line = driver.find_element(By.XPATH, f"//div[text()='{fruit}']/ancestor::div[@role='row']")
price = fruit_line.find_element(By.XPATH, ".//div[@id='cell-4-undefined']")
assert price.text == str(new_price), f'Price is not correct: {price.text}'

print('Test passed!')
# Close the browser
driver.quit()